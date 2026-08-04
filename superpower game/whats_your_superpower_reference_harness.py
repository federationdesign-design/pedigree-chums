"""Reference harness, configuration MVP-4.1, final approved name set."""
import re, json, itertools, hashlib
from collections import Counter
from openpyxl import load_workbook
POWERS=['Focus','Vision','Zoom','Ideas','Energy']
CONFIG='MVP-4.1'; SCHEMA='result-contract-1.0'
exec(open('final_names.py').read().split("def toks")[0].split("print(")[0])
wb=load_workbook('/mnt/user-data/uploads/whats_your_superpower_question_bank_v4_0.xlsx',read_only=True)
Q={}
for r in wb['Scoring Map'].iter_rows(values_only=True):
    if r[0] and re.fullmatch(r'M[0-9]{2}',str(r[0]).strip()):
        Q.setdefault(str(r[0]).strip(),{})[str(r[1]).strip()]=(r[3],r[4],r[5],r[6])
QIDS=sorted(Q)

def resolve(arr):
    raw={p:0 for p in POWERS}; prim={p:0 for p in POWERS}
    for q,s in zip(QIDS,arr):
        pp,pv,sp,sv=Q[q][s]; raw[pp]+=pv; raw[sp]+=sv; prim[pp]+=1
    top=max(raw.values()); lead=[p for p in POWERS if raw[p]==top]; n=len(lead)
    def sup_of(c):
        if len(c)==1: return c[0]
        b=max(prim[p] for p in c); c=[p for p in c if prim[p]==b]
        return sorted(c,key=POWERS.index)[0]
    if n<=2:
        rest=[p for p in POWERS if p not in lead]; sec=max(raw[p] for p in rest)
        sup=sup_of([p for p in rest if raw[p]==sec])
    else: sup=None
    if n==1:
        gap=top-sec; st='SINGLE_CLOSE' if gap<=3 else 'SINGLE_CLEAR'
        title,line=D[(lead[0],sup)]
        summ=('Your answers leaned most towards %s, with %s close behind.'%(lead[0],sup) if gap<=3
              else 'Your answers leaned most towards %s, with %s next.'%(lead[0],sup))
        tk='COMBO:%s>%s'%(lead[0],sup)
    elif n==2:
        st='TIE_TWO'; title,line=L[frozenset(lead)]
        summ='Your answers split between %s and %s, with %s next.'%(lead[0],lead[1],sup)
        tk='LEVEL:'+'+'.join(sorted(lead))
    else:
        st={3:'TIE_THREE',4:'TIE_FOUR',5:'TIE_FIVE'}[n]; title='The Power Pack'; line=None
        summ=('Your answers divided evenly across all five powers this round.' if n==5 else
              'Your answers split across '+', '.join(lead[:-1])+' and '+lead[-1]+'.')
        tk='GENERIC:power_pack'
    return st,lead,sup,title,line,summ,tk

states=Counter(); rendered=Counter(); recs=[]; used=Counter()
for arr in itertools.product('AB',repeat=15):
    st,lead,sup,title,line,summ,tk=resolve(arr)
    states[st]+=1; used[tk]+=1
    rendered[title+' | '+summ]+=1
    recs.append(json.dumps({'answerPattern':''.join(arr),'stateId':st,'leadingPowers':lead,
        'supportingPower':sup,'titleKey':tk,'summaryKey':st,
        'chartPrimaryEmphasisSet':lead,'chartSecondaryEmphasisSet':[sup] if sup else []},
        sort_keys=True,separators=(',',':')))
N=len(recs)
header=f'config={CONFIG};schema={SCHEMA};arrays={N};nameset=final-1.0'
blob=header+'\n'+'\n'.join(recs)+'\n'
digest=hashlib.sha256(blob.encode('utf-8')).hexdigest()
exp={'SINGLE_CLOSE':22064,'SINGLE_CLEAR':3901,'TIE_TWO':5959,'TIE_THREE':785,'TIE_FOUR':39,'TIE_FIVE':20}
print('state counts match locked targets:', all(states[k]==v for k,v in exp.items()))
for k in exp: print(f'  {k:13} {states[k]:6}  {100*states[k]/N:7.3f}%')
print(f'\ndirectional titles reachable: {sum(1 for k in used if k.startswith("COMBO"))} of 20')
print(f'level titles reachable:       {sum(1 for k in used if k.startswith("LEVEL"))} of 10')
print(f'distinct rendered results:    {len(rendered)}')
print(f'\nheader:  {header}\nSHA-256: {digest}')
with open('/mnt/user-data/outputs/whats_your_superpower_golden_results_final.csv','w') as f:
    f.write('count,share_percent,title,summary\n')
    for t,c in rendered.most_common():
        ti,su=t.split(' | '); f.write('%d,%.4f,"%s","%s"\n'%(c,100*c/N,ti,su))
open('/mnt/user-data/outputs/whats_your_superpower_golden_hash_final.txt','w').write(
 "What's Your Superpower, configuration MVP-4.1, result-contract schema 1.0, name set final-1.0\n\n"
 "Canonical input: one JSON record per answer array, M01 to M15 in fixed order, answer A before\n"
 "answer B, power arrays in fixed power order, sorted object keys, UTF-8, LF, no presentation markup.\n\n"
 f"header:  {header}\nSHA-256: {digest}\n\nAudit log\n"
 "  MVP-4.0 + errata 4.0.2, main-power titles (pre-result-contract):\n"
 "  def21fc20e42201cc121353f960f2179f3587dbc61d47bdbf59001f214779559\n"
 "  MVP-4.1, result contract, provisional name set:\n"
 "  7ed0478570bb69dae75ac9cc89982597a6139256fdb80ff2341b300b47a5198e\n")
